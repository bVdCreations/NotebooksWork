# import os

import openpyxl
from openpyxl.utils import get_column_letter

toprow = {'A1': '',
                'B1': 'Display Name',
                'C1': 'Host Name',
                'D1': 'Address',
                'E1': 'Device Type',
                'F1': 'Status',
                'G1': 'nDeviceID',
                'H1': 'nParentGroupID',}


def export_to(file_name, input):

    # create a workbook
    wb = openpyxl.Workbook()

    # active sheet and name it
    wbsheet = wb.active
    wbsheet.title = "devices"

    for row in input.values():
        # loop through the input the input rows
        row_number = row['row']

        # get the rownumber
        # we add to because it start at 0 (+1) and we want to move all the rows by one (+1)
        row_str = str(row_number + 2)

        for key, value in row['col'].items():
            # loop through all the items input rows
            # {0: '10.136.38.3', 1: '10.136.38.3', 2: '10.136.38.3', 3: 'Device', 4: ''}

            # get the columletter
            # we add to because it start at 0 (+1) and we want to move all the colunms by one (+1)
            col = get_column_letter(key + 2)

            # write the text in the given cell
            wbsheet[col + row_str] = value

        # add the rownumber at the first column
        wbsheet['A' + row_str] = row_number

        wbsheet['G' + row_str] = row['nDeviceID']

        wbsheet['H' + row_str] = row['nParentGroupID']

        # add the heading descripted in toprow
        for keys, values in toprow.items():
            wbsheet[keys] = values

    # save the file
    wb.save(check_filename(file_name))


def check_filename(file_name):
    return file_name
